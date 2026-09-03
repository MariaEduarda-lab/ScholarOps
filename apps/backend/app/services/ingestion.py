import csv
import io
from collections import defaultdict
from datetime import UTC, date, datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import PROJECT_DIR
from app.models import Candidate, Document, IngestionJob, Institution

INSTITUTIONS = {
    "inteli": {
        "name": "Instituto de Tecnologia e Liderança",
        "short_name": "Inteli",
        "initials": "IN",
        "process_name": "Inteli Social",
        "edition": "Graduação 2026",
        "education_level": "Ensino superior",
        "description": "Apoio integral ou parcial para estudantes da graduação, conforme avaliação socioeconômica.",
    },
    "bom_aluno_bh": {
        "name": "Instituto Bom Aluno de Belo Horizonte",
        "short_name": "Bom Aluno BH",
        "initials": "BA",
        "process_name": "Seleção Bom Aluno",
        "edition": "Seleção 2026",
        "education_level": "Educação básica",
        "description": "Identificação e acompanhamento de estudantes com critérios acadêmicos e socioeconômicos.",
    },
    "marista_dom_silverio": {
        "name": "Colégio Marista Dom Silvério",
        "short_name": "Marista Dom Silvério",
        "initials": "MD",
        "process_name": "Bolsa Social",
        "edition": "Bolsa Social 2026",
        "education_level": "Educação básica",
        "description": "Concessão de bolsas sociais mediante análise documental e avaliação socioeconômica.",
    },
}
FIRST_NAMES = [
    "Ana",
    "Bruno",
    "Camila",
    "Daniel",
    "Elisa",
    "Felipe",
    "Gabriela",
    "Henrique",
    "Isabela",
    "João",
    "Larissa",
    "Mateus",
    "Natália",
    "Otávio",
    "Paula",
    "Rafael",
    "Sofia",
    "Thiago",
    "Vitória",
    "Yasmin",
]
LAST_NAMES = [
    "Almeida",
    "Barbosa",
    "Cardoso",
    "Dias",
    "Ferreira",
    "Gomes",
    "Lima",
    "Martins",
    "Nascimento",
    "Oliveira",
    "Pereira",
    "Ramos",
    "Rocha",
    "Santos",
    "Silva",
    "Souza",
]


def labelize(value: Any) -> str:
    return str(value or "").replace("_", " ").title()


def safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def safe_float(value: Any, default: float = 0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def ensure_institutions(db: Session) -> None:
    for institution_id, payload in INSTITUTIONS.items():
        institution = db.get(Institution, institution_id)
        if institution is None:
            db.add(Institution(id=institution_id, **payload))
        else:
            for field, value in payload.items():
                setattr(institution, field, value)
    db.commit()


def parse_csv_bytes(contents: bytes) -> list[dict[str, str]]:
    try:
        text = contents.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("O CSV precisa estar codificado em UTF-8.") from exc
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("O CSV não possui cabeçalho.")
    return [dict(row) for row in reader]


def parse_xlsx_bytes(contents: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            raise ValueError("A planilha XLSX não possui uma aba ativa.")
        raw_rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(raw_rows)]
    except (OSError, ValueError, StopIteration) as exc:
        raise ValueError("A planilha XLSX está vazia ou não pôde ser lida.") from exc
    if not all(headers):
        raise ValueError("Todas as colunas da planilha precisam ter um cabeçalho.")
    return [dict(zip(headers, row, strict=True)) for row in raw_rows if any(value is not None for value in row)]


def normalize_document_status(row: dict[str, Any]) -> str:
    raw = str(row.get("status_documento", "")).lower()
    issue = str(row.get("pendencia_esperada", "")).lower()
    if "ileg" in raw or "ileg" in issue:
        return "ilegivel"
    if "incons" in raw or "diverg" in issue or "incons" in issue:
        return "inconsistente"
    if raw == "ok" and issue in {"nenhuma", "nao_se_aplica", ""}:
        return "ok"
    return "pendente"


def candidate_name(external_id: str) -> str:
    seed = int("".join(character for character in external_id if character.isdigit()) or "1")
    first = FIRST_NAMES[(seed * 7) % len(FIRST_NAMES)]
    last = LAST_NAMES[(seed * 11) % len(LAST_NAMES)]
    second = LAST_NAMES[(seed * 3 + 5) % len(LAST_NAMES)]
    return f"{first} {last} {second}"


def _candidate_from_rows(
    db: Session, institution_id: str, external_id: str, rows: list[dict[str, Any]]
) -> tuple[Candidate, int]:
    base = rows[0]
    candidate_pk = f"{institution_id}:{external_id}"
    candidate = db.get(Candidate, candidate_pk)
    if candidate is None:
        seed = safe_int("".join(character for character in external_id if character.isdigit()) or "1", 1)
        name = str(base.get("nome_candidato") or base.get("nome") or candidate_name(external_id))
        members = len({str(row.get("membro_id", "candidato")) for row in rows})
        income = safe_float(base.get("renda_familiar"), 1350 + ((seed * 337) % 4650))
        phone_digits = str(80_000_000 + ((seed * 7919) % 9_999_999)).zfill(8)
        candidate = Candidate(
            id=candidate_pk,
            external_id=external_id,
            institution_id=institution_id,
            name=name,
            age=safe_int(base.get("idade")),
            edition=labelize(base.get("edicao", "")),
            scenario=labelize(base.get("cenario_teste", "")),
            submitted_at=date(2026, 4 + (seed % 2), 2 + (seed % 25)),
            phone=f"(31) 9{phone_digits[:4]}-{phone_digits[4:]}",
            email=f"candidato.{seed}@email.com",
            city="São Paulo, SP" if institution_id == "inteli" else "Belo Horizonte, MG",
            family_members=max(members, 1),
            monthly_income=income,
            per_capita_income=round(income / max(members, 1), 2),
        )
        db.add(candidate)
        db.flush()
    else:
        candidate.edition = labelize(base.get("edicao", candidate.edition))
        candidate.scenario = labelize(base.get("cenario_teste", candidate.scenario))

    documents_upserted = 0
    for index, row in enumerate(rows, start=1):
        external_document_id = str(row.get("arquivo_id") or f"{external_id}-DOC-{index:04d}")
        document_id = f"{institution_id}:{external_id}:{external_document_id}"
        document = db.get(Document, document_id)
        if document is None:
            document = Document(id=document_id, candidate_id=candidate.id, institution_id=institution_id)
            db.add(document)
        pending_raw = str(row.get("pendencia_esperada") or "")
        document.category = labelize(row.get("categoria_documental"))
        document.document_type = str(row.get("tipo_documento") or "documento")
        document.label = labelize(document.document_type)
        document.member_id = str(row.get("membro_id") or "candidato")
        document.relationship_label = labelize(row.get("relacao") or "candidato")
        document.required = labelize(row.get("obrigatoriedade") or "a_validar")
        document.status = normalize_document_status(row)
        document.issue = None if pending_raw in {"", "nenhuma", "nao_se_aplica"} else labelize(pending_raw)
        document.confidence = safe_float(row.get("confianca_extracao"))
        document.declared_value = str(row.get("valor_declarado") or "")
        document.extracted_value = str(row.get("valor_extraido") or "")
        document.human_review = str(row.get("revisao_humana_esperada") or "").lower() == "sim"
        document.source_payload = dict(row)
        documents_upserted += 1
    db.flush()
    db.refresh(candidate)
    recompute_candidate(candidate)
    return candidate, documents_upserted


def recompute_candidate(candidate: Candidate) -> None:
    documents = candidate.documents
    candidate.pending_count = sum(document.status == "pendente" for document in documents)
    candidate.inconsistent_count = sum(document.status == "inconsistente" for document in documents)
    candidate.attention_count = sum(document.human_review or document.status == "ilegivel" for document in documents)
    ok_count = sum(document.status == "ok" for document in documents)
    candidate.progress = round((ok_count / len(documents)) * 100) if documents else 0
    if candidate.inconsistent_count:
        candidate.status = "Em revisão"
    elif candidate.pending_count:
        candidate.status = "Documentação pendente"
    elif candidate.attention_count:
        candidate.status = "Aguardando análise"
    else:
        candidate.status = "Apto para entrevista"
    candidate.insights = [
        f"{candidate.pending_count} documento(s) ainda exigem complementação."
        if candidate.pending_count
        else "Documentos obrigatórios recebidos.",
        f"{candidate.inconsistent_count} divergência(s) identificada(s)."
        if candidate.inconsistent_count
        else "Não foram identificadas divergências automáticas.",
        f"{candidate.attention_count} item(ns) precisam de interpretação humana."
        if candidate.attention_count
        else "Extrações com boa confiança para triagem.",
    ]
    formatted_income = f"{candidate.monthly_income:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    candidate.summary = (
        f"{candidate.name} possui {candidate.family_members} pessoa(s) no núcleo familiar e renda mensal "
        f"informada de R$ {formatted_income}. A triagem organizou {len(documents)} documento(s) "
        "e destacou evidências que devem ser conferidas antes da entrevista."
    )
    candidate.updated_at = datetime.now(UTC)


def ingest_rows(
    db: Session,
    institution_id: str,
    rows: list[dict[str, Any]],
    source_type: str,
    filename: str | None = None,
) -> IngestionJob:
    job = IngestionJob(
        institution_id=institution_id,
        source_type=source_type,
        filename=filename,
        rows_received=len(rows),
    )
    db.add(job)
    db.flush()
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    errors: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        row_institution = str(row.get("instituicao") or institution_id)
        if row_institution != institution_id:
            errors.append({"row": row_number, "message": "A linha pertence a outra instituição."})
            continue
        external_id = str(row.get("candidatura_id") or row.get("candidate_id") or "").strip()
        if not external_id:
            errors.append({"row": row_number, "message": "candidatura_id é obrigatório."})
            continue
        grouped[external_id].append(row)

    documents_upserted = 0
    for external_id, candidate_rows in grouped.items():
        _, count = _candidate_from_rows(db, institution_id, external_id, candidate_rows)
        documents_upserted += count
    job.candidates_upserted = len(grouped)
    job.documents_upserted = documents_upserted
    job.errors = errors[:100]
    job.status = "completed_with_errors" if errors else "completed"
    job.completed_at = datetime.now(UTC)
    db.commit()
    db.refresh(job)
    return job


def seed_demo_database(db: Session) -> None:
    ensure_institutions(db)
    if db.scalar(select(Candidate.id).limit(1)) is not None:
        return
    files = sorted((PROJECT_DIR / "dados" / "sinteticos").glob("*_documentos_sinteticos.csv"))
    for file_path in files:
        rows = parse_csv_bytes(file_path.read_bytes())
        if not rows:
            continue
        institution_id = str(rows[0].get("instituicao") or "")
        if institution_id in INSTITUTIONS:
            ingest_rows(db, institution_id, rows, source_type="seed", filename=file_path.name)
